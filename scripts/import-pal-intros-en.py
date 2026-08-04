import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from openpyxl import load_workbook

workbook_path = sys.argv[1]
output_path = sys.argv[2]
with open("public/data/pals.json", encoding="utf-8") as handle:
    pals = json.load(handle)

pal_by_name = {pal["name"].replace(" ", "_"): pal for pal in pals}
sheet = load_workbook(workbook_path, read_only=True, data_only=True)["帕鲁数据（299）"]
rows = list(sheet.iter_rows(min_row=5, values_only=True))
rows = [row for row in rows if row and row[0] and row[15]]
if len(rows) != 299:
    raise ValueError(f"Expected 299 intro rows, found {len(rows)}")

def translate(text):
    query = urllib.parse.urlencode({"client": "gtx", "sl": "zh-CN", "tl": "en", "dt": "t", "q": text})
    request = urllib.request.Request(
        f"https://translate.googleapis.com/translate_a/single?{query}",
        headers={"User-Agent": "Mozilla/5.0"},
    )
    with urllib.request.urlopen(request, timeout=20) as response:
        payload = json.load(response)
    translated = " ".join(part[0] for part in payload[0] if part and part[0]).strip()
    translated = re.sub(r"\s+", " ", translated)
    replacements = {
        "the sight of Venus": "seeing stars",
        "sight of Venus": "seeing stars",
        "Mian Youyou": "Lamball",
        "cotton yoyo": "Lamball",
        "naughty cat": "Cattiva",
        "Melpa": "Melpaca",
        "Paros Islands": "Palpagos Islands",
    }
    for source, target in replacements.items():
        translated = translated.replace(source, target)
    translated = re.sub(r"\b(?:Palu|palu|Paru|paru|Pallu|pallu|Parlu|parlu|Palluyong|palluyong)\b", "Pal", translated)
    return translated

partial_path = f"{output_path}.partial"
if os.path.exists(partial_path):
    with open(partial_path, encoding="utf-8") as handle:
        output = json.load(handle)
else:
    output = []

for index, row in enumerate(rows[len(output):], start=len(output) + 1):
    name, intro = row[0], row[15]
    pal = pal_by_name.get(str(name).replace(" ", "_"))
    if not pal:
        raise ValueError(f"No matching Pal record for {name}")
    for attempt in range(5):
        try:
            translated = translate(str(intro))
            break
        except Exception:
            if attempt == 4:
                raise
            time.sleep(2 ** attempt)
    output.append({"palSlug": pal["slug"], "name": pal["name"], "intro": translated})
    with open(partial_path, "w", encoding="utf-8") as handle:
        json.dump(output, handle, ensure_ascii=False)
    print(f"{index}/299 {name}", flush=True)
    time.sleep(0.15)

with open(output_path, "w", encoding="utf-8") as handle:
    json.dump(output, handle, ensure_ascii=False, indent=2)
    handle.write("\n")
os.remove(partial_path)
print(f"Wrote {len(output)} English Pal introductions.")
